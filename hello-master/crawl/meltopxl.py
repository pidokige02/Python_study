import csv
import codecs
import openpyxl
from pprint import pprint
from openpyxl.chart import (
    Reference,
    BarChart,
    Series,
    ScatterChart
)

csvFile = codecs.open("../crawl/data/meltop100.csv", "r", "utf-8")
reader = csv.reader(csvFile, delimiter=',', quotechar='"')

book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = "문제1"

for i, row in enumerate(reader):
    for j, cols in enumerate(row):
        val = row[j]
        if (j == 0 or j == 3 or j == 4) and val.isnumeric():
            val = int(row[j]) 

        sheet1.cell(row=(i+1), column=(j+1)).value = val

sheet2 = book.create_sheet()
sheet2.title = "문제2"

for i in range(100):
    newFile = './images/' + str(i+1) + '_50x50.jpg'
    img3 = openpyxl.drawing.image.Image(newFile)
    sheet2.row_dimensions[(i+1)].height = 42
    sheet2.add_image(img3, 'A' + str(i+1))

sheet3 = book.create_sheet()
sheet3.title = "문제3"

datax = Reference(sheet1, min_col=4,
                  min_row=2, max_col=4, max_row=11)
categs = Reference(sheet1, min_col=1,
                   min_row=2, max_row=11)

chart = BarChart()
chart.add_data(data=datax)
chart.set_categories(categs)

chart.legend = None  # 범례
chart.varyColors = True
chart.title = "Top10 좋아요"

sheet3.add_chart(chart, "B2")

## ------------------------------ scatter
chart = ScatterChart()
chart.style = 13
chart.x_axis.title = 'Size'
chart.y_axis.title = 'Percentage'

xvalues = Reference(sheet1, min_col=1,
                    min_row=2, max_row=11)

values = Reference(sheet1,
                    min_col=5,
                    min_row=2,
                    max_row=11)
series = Series(values, xvalues,
                title_from_data=True)
chart.series.append(series)

sheet3.add_chart(chart, "B20")


book.save('./data/meltop100.xlsx')
