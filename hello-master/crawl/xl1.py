from PIL import Image
import openpyxl
import datetime
from openpyxl.styles.borders import Border, Side
from openpyxl.chart import (
    Reference, Series,
    ScatterChart
)

book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = "첫번째 시트"
sheet1.cell(row=1, column=1).value = 'Title'
sheet2 = book.create_sheet()
sheet2.title = "두번째 시트"
sheet2['A1'] = datetime.datetime.now()
sheet2['A2'] = datetime.date.today()
sheet2['C1'] = '121'

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='none'),
                     bottom=Side(style='thin'))

cell = sheet2.cell(row=3, column=2)
cell.value = "Border!!"
cell.border = thin_border

imgFile = '../crawl/images/aaa.png'
img = openpyxl.drawing.image.Image(imgFile)
sheet1.add_image(img, 'B5')

img2 = Image.open(imgFile)
new_img = img2.resize((100, 100))
new_img.save('new.png')
img3 = openpyxl.drawing.image.Image('new.png')
sheet2.add_image(img3, 'D1')


rows = [
    ['Size', 'Batch 1', 'Batch 2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 25],
    [6, 25, 35],
    [7, 20, 40],
]

sheet3 = book.create_sheet()
for row in rows:
    sheet3.append(row)

chart = ScatterChart()
chart.style = 13
chart.x_axis.title = 'Size'
chart.y_axis.title = 'Percentage'

xvalues = Reference(sheet3, min_col=1,
                    min_row=2, max_row=7)

for i in range(2, 4):
    values = Reference(sheet3,
                       min_col=i,
                       min_row=1,
                       max_row=7)
    series = Series(values, xvalues,
                    title_from_data=True)
    chart.series.append(series)

sheet3.add_chart(chart, "A10")


# 저장하기
book.save("./data/output.xlsx")
