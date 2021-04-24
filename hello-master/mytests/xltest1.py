import csv, codecs
import openpyxl
from pprint import pprint
import datetime

csvFile = codecs.open("../crawl/data/meltop100.csv", "r", "utf-8")
reader = csv.reader(csvFile, delimiter=',', quotechar='"')

book = openpyxl.Workbook()
sheet1 = book.active

sheet1.title = "첫번째 시트"

for i, row in enumerate(reader):
    # print(i, row)
    for j, col in enumerate(row):
        tcell = sheet1.cell(row=(i+1), column=j+1)
        if i > 0 and (j == 0 or j > 2) and col.isnumeric():
            print(j, col)
            tcell.number_format
            tcell.value = int(col)
        else:
            tcell.value = col

sheet2 = book.create_sheet()
sheet2.title = "두번째 시트"
sheet2['A1'] = datetime.datetime.now()
sheet2['A2'] = datetime.date.today()
sheet2['B1'] = '하하하'

tmpCell = sheet2['C1']
tmpCell.font = openpyxl.styles.Font(size=14, color='FF0000')
tmpCell.number_format
tmpCell.value = 12345

book.save("t1.xlsx")
exit()