from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy, deepcopy
import sys

new_file_name = "osprey_issues_new.xlsx"
old_file_name = "osprey_issues_old.xlsx"

if len(sys.argv) == 1:
    txt = "running : Python xl_smart.py {} {}".format(old_file_name, new_file_name)
    print(txt) 
elif len(sys.argv) == 2:
    print("Usage : Python xl_smart.py oldfilename newfilename!")
    txt = "running : Python xl_smart.py {} {}".format(old_file_name, new_file_name)
    print(txt) 
    sys.exit()
elif len(sys.argv) == 3:
    old_file_name = sys.argv[1]
    new_file_name = sys.argv[2]
    txt = "running : Python xl_smart.py {} {}".format(old_file_name, new_file_name)
    print(txt) 
else:  
    print("Usage : Python xl_smart.py oldfilename newfilename!")
    sys.exit()

wb_old = load_workbook(old_file_name)
wb_new = load_workbook(new_file_name)

print(old_file_name)
print(new_file_name)

print(wb_new.sheetnames)
print(wb_old.sheetnames)


ws_new = wb_new[wb_new.sheetnames[0]]  #latest weeks SPR
ws_previous = wb_old[wb_old.sheetnames[0]] #previous weeks SPR

print(ws_new)
print(ws_previous)

for row in ws_previous.iter_rows(min_row=1, min_col=27, max_col=29):
    for cell in row:
        ws_new[cell.coordinate].value = ws_previous[cell.coordinate].value
        ws_new[cell.coordinate].font = copy(ws_previous[cell.coordinate].font)
        ws_new[cell.coordinate].border = copy(ws_previous[cell.coordinate].border)
        ws_new[cell.coordinate].fill = copy(ws_previous[cell.coordinate].fill)
        ws_new[cell.coordinate].number_format = copy(ws_previous[cell.coordinate].number_format)
        ws_new[cell.coordinate].protection = copy(ws_previous[cell.coordinate].protection)
        ws_new[cell.coordinate].alignment = copy(ws_previous[cell.coordinate].alignment)
   
# Comment, highlight column
ws_new.column_dimensions[get_column_letter(27)].width = ws_previous.column_dimensions[get_column_letter(27)].width    
ws_new.column_dimensions[get_column_letter(28)].width = ws_previous.column_dimensions[get_column_letter(28)].width    
ws_new.column_dimensions[get_column_letter(29)].width = ws_previous.column_dimensions[get_column_letter(29)].width    

# change sheet name of new comsolidated file
position = new_file_name.find('FW')
# print ('position', position)
# print ('slice', new_file_name[position:position+4])

ws_new.title = new_file_name[position:position+4] # slice FWXX and change sheet title w/ it

wb_new.save(new_file_name)
wb_old.save(old_file_name)
wb_new.close()
wb_old.close()

