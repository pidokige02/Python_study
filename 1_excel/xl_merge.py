from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy, deepcopy
import sys

new_file_name = "osprey_issues_new.xlsx"
old_file_name = "osprey_issues_old.xlsx"
txt = ""

if len(sys.argv) == 1:
    txt = "running : Python xl_merge.py {} {}".format(old_file_name, new_file_name)
    print(txt) 
elif len(sys.argv) == 2:
    print("Usage : Python xl_merge.py oldfilename newfilename!")
    txt = "running : Python xl_merge.py {} {}".format(old_file_name, new_file_name)
    print(txt) 
    sys.exit()
elif len(sys.argv) == 3:
    old_file_name = sys.argv[1]
    new_file_name = sys.argv[2]
    txt = "running : Python xl_merge.py {} {}".format(old_file_name, new_file_name)
    print(txt) 
else:  
    print("Usage : Python xl_merge.py oldfilename newfilename!")
    sys.exit()

wb_old = load_workbook(old_file_name)
wb_new = load_workbook(new_file_name)

ws_olds = []  # using list 
ws_news = []  # using list 
flag_for_ws_olds = {} # using dictionalry type

print(wb_old.sheetnames)
print(wb_new.sheetnames)

# storing the information for ws creation in target wb
for sheetname in wb_old.sheetnames:
    brank = sheetname.find("Sheet")
    if(brank != -1): # blank Sheet exists
        print("blank Sheet exists and continue")
        continue
    for new_sheetname in wb_new.sheetnames:
        duplicate = new_sheetname.find(sheetname)      
        if( duplicate != -1): #  duplicate exists  
            flag_for_ws_olds[sheetname] = False
            break
        else:
            flag_for_ws_olds[sheetname] = True


print(flag_for_ws_olds)

# adding new ws only, not already duplicated ones
for key in flag_for_ws_olds:
    val = flag_for_ws_olds[key]
    if (val == True):
        ws_olds.append(wb_old[key])
        ws_news.append(wb_new.create_sheet(key))

if len(ws_olds) == 0:
    print("No data to Process")
    sys.exit()

index = 0
for ws_old in ws_olds:
    for row in ws_old:
        for cell in row:
            ws_news[index][cell.coordinate].value = cell.value
            ws_news[index][cell.coordinate].font = copy(cell.font)
            ws_news[index][cell.coordinate].border = copy(cell.border)
            ws_news[index][cell.coordinate].fill = copy(cell.fill)
            ws_news[index][cell.coordinate].number_format = copy(cell.number_format)
            ws_news[index][cell.coordinate].protection = copy(cell.protection)
            ws_news[index][cell.coordinate].alignment = copy(cell.alignment)
    
    for i in range(1, ws_news[index].max_column+1):
        ws_news[index].column_dimensions[get_column_letter(i)].width = ws_old.column_dimensions[get_column_letter(i)].width    

    index += 1


ws_final = wb_new[wb_new.sheetnames[0]]

index = 0
for sheetname in wb_new.sheetnames:
    if(sheetname.find("Summary") == -1): # if Summary sheet not found
        index+=1
    else:       # if Summary sheet  found
        break 

if (index != 0):
    print("-index", (-1)*index)
    wb_new.move_sheet("Summary",(-1)*index )  # move summary sheet first

ws_final = wb_new[wb_new.sheetnames[1]]  #latest weeks SPR
ws_previous = wb_new[wb_new.sheetnames[2]] #previous weeks SPR

print(ws_final)
print(ws_previous)

for row in ws_previous.iter_rows(min_row=1, min_col=27, max_col=29):
    for cell in row:
        ws_final[cell.coordinate].value = ws_previous[cell.coordinate].value
        ws_final[cell.coordinate].font = copy(ws_previous[cell.coordinate].font)
        ws_final[cell.coordinate].border = copy(ws_previous[cell.coordinate].border)
        ws_final[cell.coordinate].fill = copy(ws_previous[cell.coordinate].fill)
        ws_final[cell.coordinate].number_format = copy(ws_previous[cell.coordinate].number_format)
        ws_final[cell.coordinate].protection = copy(ws_previous[cell.coordinate].protection)
        ws_final[cell.coordinate].alignment = copy(ws_previous[cell.coordinate].alignment)
   
# Comment, highlight column
ws_final.column_dimensions[get_column_letter(27)].width = ws_previous.column_dimensions[get_column_letter(27)].width    
ws_final.column_dimensions[get_column_letter(28)].width = ws_previous.column_dimensions[get_column_letter(28)].width    
ws_final.column_dimensions[get_column_letter(29)].width = ws_previous.column_dimensions[get_column_letter(28)].width    

wb_new.save(new_file_name)
wb_old.save(old_file_name)
wb_new.close()
wb_old.close()

# test code to copy the whole contents from legacy sheet.
# ws_old = wb_old['FW05']
# ws_new = wb_new.create_sheet('FW05')

# for row in ws_old:
#     for cell in row:
#         ws_new[cell.coordinate].value = cell.value
#         ws_new[cell.coordinate].font = copy(cell.font)
#         ws_new[cell.coordinate].border = copy(cell.border)
#         ws_new[cell.coordinate].fill = copy(cell.fill)
#         ws_new[cell.coordinate].number_format = copy(cell.number_format)
#         ws_new[cell.coordinate].protection = copy(cell.protection)
#         ws_new[cell.coordinate].alignment = copy(cell.alignment)

# for i in range(1, ws_new.max_column+1):
#     ws_new.column_dimensions[get_column_letter(i)].width = ws_old.column_dimensions[get_column_letter(i)].width    

        
# print(ws_old)
