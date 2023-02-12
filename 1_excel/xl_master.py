from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy, deepcopy
import sys

add_file_name = "osprey_issues_FW07.xlsx"
master_file_name = "osprey_issues_master.xlsx"

if len(sys.argv) == 1:
    txt = "running : Python xl_master.py {}".format(add_file_name)
    print(txt) 
    sys.exit()
elif len(sys.argv) == 2:
    add_file_name = sys.argv[1]
    txt = "running : Python xl_master.py {}".format(add_file_name)
    print(txt) 
else:  
    print("Usage : Python xl_master.py addfilename!")
    sys.exit()

wb_master = load_workbook(master_file_name)
wb_add = load_workbook(add_file_name)

ws_masters = []  # using list 
ws_adds = []  # using list 

print(wb_master.sheetnames)
print(wb_add.sheetnames)

find = -1 # not found
for sheetname in wb_master.sheetnames:
    find = sheetname.find(wb_add.sheetnames[0])
    if(find != -1): # duplicated Sheet exists
        print("duplicated Sheet exists in master file")
        sys.exit()
        break

ws_masters.append(wb_master.create_sheet(wb_add.sheetnames[0]))
ws_adds.append(wb_add[wb_add.sheetnames[0]])

index = 0
for ws_add in ws_adds:
    for row in ws_add:
        for cell in row:
            ws_masters[index][cell.coordinate].value = cell.value
            ws_masters[index][cell.coordinate].font = copy(cell.font)
            ws_masters[index][cell.coordinate].border = copy(cell.border)
            ws_masters[index][cell.coordinate].fill = copy(cell.fill)
            ws_masters[index][cell.coordinate].number_format = copy(cell.number_format)
            ws_masters[index][cell.coordinate].protection = copy(cell.protection)
            ws_masters[index][cell.coordinate].alignment = copy(cell.alignment)
    
    for i in range(1, ws_masters[index].max_column+1):
        ws_masters[index].column_dimensions[get_column_letter(i)].width = ws_add.column_dimensions[get_column_letter(i)].width    

    index += 1


index = 0
for sheetname in wb_master.sheetnames:
    if(sheetname.find(wb_add.sheetnames[0]) == -1): # if added sheet is not found
        index+=1
    else:       # if added sheet is found
        break 

if (index != 0):
    print("-index", (-1)*index)
    wb_master.move_sheet(wb_add.sheetnames[0],(-1)*(index-1) )  # move added sheet next to summary sheet

wb_master.save(master_file_name)
wb_add.save(add_file_name)
wb_master.close()
wb_add.close()
