from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy, deepcopy
import sys
import re

def get_position (base_position, index):
    digit =  re.findall("[0-9]+", base_position)
    character = re.findall("[A-Z]+", base_position)
    ascii_code = ord(character[0])

    new_ascii_code = ascii_code + 2 * index
    new_character = chr(new_ascii_code)
    new_position = new_character[0] + digit[0]
    return new_position

def get_delta_position (base_position, index):
    digit =  re.findall("[0-9]+", base_position)
    character = re.findall("[A-Z]+", base_position)
    ascii_code = ord(character[0])

    delta_ascii_code = ascii_code - index
    prev_ascii_code = ascii_code - 2* index

    delta_character = chr(delta_ascii_code)
    prev_character = chr(prev_ascii_code)

    delta_position = delta_character[0] + digit[0]
    prev_position = prev_character[0] + digit[0]

    return [delta_position,prev_position] 

def update_delta(ws_summary, new_position, index):
    return_position = get_delta_position (new_position, index)
    txt = "={0}-{1}".format(new_position, return_position[1]) 
    ws_summary[return_position[0]] = txt


file_name = "osprey_issues_master.xlsx"
txt = ""

total_spr_table = {
    'total' : 'B2',
    'total_NC' : 'B3',
    'total_NC_No_DRB' : 'B4',
    'total_NC_Major' : 'B5',
    'total_NC_Minor' : 'B6',
    'total_IO' : 'B7',
    'Not Started' : 'B8',
    'Not Started_NC' : 'B9',
    'Not Started_NC_No_DRB' : 'B10',
    'Not Started_NC_Major' : 'B11',
    'Not Started_NC_Minor' : 'B12',
    'Not Started_IO' : 'B13',
    'in Prog' : 'B14',
    'in Prog_NC' : 'B15',
    'in Prog_NC_No_DRB' : 'B16',
    'in Prog_NC_Major' : 'B17',
    'in Prog_NC_Minor' : 'B18',
    'in Prog_IO' : 'B19',
    'Resolved' : 'B20',
    'Verified' : 'B21',
    'Closed' : 'B22',
}


sw_spr_table = {
    'SW_total' : 'B27',
    'SW_total_NC' : 'B28',
    'SW_total_NC_No_DRB' : 'B29',
    'SW_total_NC_Major' : 'B30',
    'SW_total_NC_Minor' : 'B31',
    'SW_total_IO' : 'B32',
    'SW_Not Started' : 'B33',
    'SW_Not Started_NC' : 'B34',
    'SW_Not Started_NC_No_DRB' : 'B35',
    'SW_Not Started_NC_Major' : 'B36',
    'SW_Not Started_NC_Minor' : 'B37',
    'SW_Not Started_IO' : 'B38',
    'SW_in Prog' : 'B39',
    'SW_in Prog_NC' : 'B40',
    'SW_in Prog_NC_No_DRB' : 'B41',
    'SW_in Prog_NC_Major' : 'B42',
    'SW_in Prog_NC_Minor' : 'B43',
    'SW_in Prog_IO' : 'B44',
    'SW_Resolved' : 'B45',
    'SW_Verified' : 'B46',
    'SW_Closed' : 'B47',
}


system_spr_table = {
    'system_total' : 'B53',
    'system_total_NC' : 'B54',
    'system_total_NC_No_DRB' : 'B55',
    'system_total_NC_Major' : 'B56',
    'system_total_NC_Minor' : 'B57',
    'system_total_IO' : 'B58',
    'system_Not Started' : 'B59',
    'system_Not Started_NC' : 'B60',
    'system_Not Started_NC_No_DRB' : 'B61',
    'system_Not Started_NC_Major' : 'B62',
    'system_Not Started_NC_Minor' : 'B63',
    'system_Not Started_IO' : 'B64',
    'system_in Prog' : 'B65',
    'system_in Prog_NC' : 'B66',
    'system_in Prog_NC_No_DRB' : 'B67',
    'system_in Prog_NC_Major' : 'B68',
    'system_in Prog_NC_Minor' : 'B69',
    'system_in Prog_IO' : 'B70',
    'system_Resolved' : 'B71',
    'system_Verified' : 'B72',
    'system_Closed' : 'B73',
}

sw_members = [
        "\"=dongyoung.choi\"",
        "\"=taeyang.an\"",
        "\"=han il.lee\"",
        "\"=ho.lee\"",
        "\"=youjin.jung\"",
        "\"=doo je.sung\"",
        "\"=jin ha.hwang\"",
        "\"=youngdug.kim\"",
        "\"=youjin.na\"",
        "\"=kyudong.kim\"",
        "\"=heejin.na\""
]

system_members = [
        "\"=jihye.han\"",
        "\"=bong hyo.han\"",
        "\"=jae ha.hyun\"",
        "\"=jong gun.lee\"",
        "\"=jungho.kim\"",
        "\"=taeyun.kim\"",
        "\"=yasuhiro.yamada\"",
        "\"=dongwoo.lee\""
]


if len(sys.argv) == 1:
    txt = "running : Python xl_analysis.py {} ".format(file_name)
    print(txt) 
elif len(sys.argv) == 2:
    file_name = sys.argv[1]
    txt = "running : Python xl_analysis.py {} ".format(file_name)
    print(txt) 
else:  
    print("Usage : Python xl_analysis.py file_name.xlsx")
    sys.exit()

wb_src = load_workbook(file_name)
ws_srcs = []  # using list 
flag_for_analysis = {} # using dictionalry type

print(wb_src.sheetnames)

for sheetname in wb_src.sheetnames:
    validsheet = sheetname.find("FW")
    if(validsheet != -1):  # valid weekly spr worksheet exists
        flag_for_analysis[sheetname] = True
    else:
        flag_for_analysis[sheetname] = False

# getting target worksheet for analysis
ws_summary = wb_src[wb_src.sheetnames[0]]

# adding weekly base SPR worksheet only
for key in flag_for_analysis:
    val = flag_for_analysis[key]
    if (val == True):
        ws_srcs.append(wb_src[key])


print(ws_srcs)
print(ws_summary)

index = 0
ascii = 0
for ws_src in reversed(ws_srcs):

    new_position = get_position (total_spr_table['total'], index)
    txt = "=COUNTA('{}'!A2:A3000)".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)
             

    #NC from ALL SPR
    new_position = get_position (total_spr_table['total_NC'], index)
    txt = "=COUNTIF('{0}'!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #NC No_DRB from ALL SPR
    new_position = get_position (total_spr_table['total_NC_No_DRB'], index)
    txt = "=COUNTIFS('{0}'!J2:J3000,\"=NC - Design Non-Conformance\", '{0}'!AC2:AC3000, \"=No DRB\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #NC Major from ALL SPR
    new_position = get_position (total_spr_table['total_NC_Major'], index)
    txt = "=COUNTIFS('{0}'!J2:J3000,\"=NC - Design Non-Conformance\", '{0}'!AC2:AC3000, \"=Major\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #NC Minor from ALL SPR
    new_position = get_position (total_spr_table['total_NC_Minor'], index)
    txt = "=COUNTIFS('{0}'!J2:J3000,\"=NC - Design Non-Conformance\", '{0}'!AC2:AC3000, \"=Minor\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #IO from ALL SPR
    new_position = get_position (total_spr_table['total_IO'], index)
    txt = "=COUNTIFS('{0}'!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # Not Started from  ALL SPR
    new_position = get_position (total_spr_table['Not Started'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=Submitted\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Accepted\")\
    +COUNTIFS('{0}'!M2:M3000,\"=In Review\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Postponed\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #NC for Not Started
    new_position = get_position (total_spr_table['Not Started_NC'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=Submitted\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Accepted\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\")\
    +COUNTIFS('{0}'!M2:M3000,\"=In Review\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Postponed\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #NC No_DRB for Not Started
    new_position = get_position (total_spr_table['Not Started_NC_No_DRB'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=Submitted\",'{0}'!J2:J3000, \"=NC - Design Non-Conformance\",'{0}'!AC2:AC3000,\"=No DRB\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Accepted\",'{0}'!J2:J3000, \"=NC - Design Non-Conformance\",'{0}'!AC2:AC3000, \"=No DRB\")\
    +COUNTIFS('{0}'!M2:M3000,\"=In Review\",'{0}'!J2:J3000, \"=NC - Design Non-Conformance\",'{0}'!AC2:AC3000, \"=No DRB\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Postponed\",'{0}'!J2:J3000, \"=NC - Design Non-Conformance\",'{0}'!AC2:AC3000, \"=No DRB\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #NC Major for Not Started
    new_position = get_position (total_spr_table['Not Started_NC_Major'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000, \"=Submitted\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\", '{0}'!AC2:AC3000, \"=Major\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Accepted\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\",'{0}'!AC2:AC3000,\"=Major\")\
    +COUNTIFS('{0}'!M2:M3000,\"=In Review\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\",'{0}'!AC2:AC3000,\"=Major\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Postponed\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\",'{0}'!AC2:AC3000,\"=Major\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #NC Minor for Not Started
    new_position = get_position (total_spr_table['Not Started_NC_Minor'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=Submitted\", '{0}'!J2:J3000, \"=NC - Design Non-Conformance\",'{0}'!AC2:AC3000, \"=Minor\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Accepted\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\",'{0}'!AC2:AC3000, \"=Minor\")\
    +COUNTIFS('{0}'!M2:M3000,\"=In Review\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\",'{0}'!AC2:AC3000, \"=Minor\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Postponed\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\",'{0}'!AC2:AC3000, \"=Minor\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #IO for Not Strated
    new_position = get_position (total_spr_table['Not Started_IO'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=Submitted\",'{0}'!J2:J3000,\"=IO - Improvement Opportunity\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Accepted\",'{0}'!J2:J3000,\"=IO - Improvement Opportunity\")\
    +COUNTIFS('{0}'!M2:M3000,\"=In Review\",'{0}'!J2:J3000,\"=IO - Improvement Opportunity\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Postponed\",'{0}'!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # In Progress from ALL SPR
    new_position = get_position (total_spr_table['in Prog'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=In Progress\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #NC for In Progress 
    new_position = get_position (total_spr_table['in Prog_NC'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=In Progress\" ,'{0}'!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #NC No_DRB for In Progress 
    new_position = get_position (total_spr_table['in Prog_NC_No_DRB'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=In Progress\" ,'{0}'!J2:J3000,\"=NC - Design Non-Conformance\", '{0}'!AC2:AC3000,\"=No DRB\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #NC Major for In Progress 
    new_position = get_position (total_spr_table['in Prog_NC_Major'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=In Progress\" ,'{0}'!J2:J3000,\"=NC - Design Non-Conformance\", '{0}'!AC2:AC3000,\"=Major\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #NC Minor for In Progress 
    new_position = get_position (total_spr_table['in Prog_NC_Minor'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=In Progress\" ,'{0}'!J2:J3000,\"=NC - Design Non-Conformance\", '{0}'!AC2:AC3000,\"=Minor\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #IO for In Progress 
    new_position = get_position (total_spr_table['in Prog_IO'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=In Progress\" ,'{0}'!J2:J3000,\"IO - Improvement Opportunity\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #Resolved 
    new_position = get_position (total_spr_table['Resolved'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=Resolved\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #Verified 
    new_position = get_position (total_spr_table['Verified'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=Verified\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    #Closed 
    new_position = get_position (total_spr_table['Closed'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=Closed\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # # total SPRs owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ")+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_total'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # # NC SPRs owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_total_NC'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # NC NO DRB SPRs owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=No_DRB\")".format(ws_src.title) + "+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_total_NC_No_DRB'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # NC Major SPRs owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",'{0}'!AC2:AC3000,\"=Major\")".format(ws_src.title) + "+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_total_NC_Major'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # NC Minor SPRs owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",'{0}'!AC2:AC3000,\"=Minor\")".format(ws_src.title) + "+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_total_NC_Minor'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # IO  SPRs owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_total_IO'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # # Not Started SPRs owned by s/w team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    for sw_member in sw_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Submitted\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Accepted\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=In Review\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Postponed\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_Not Started'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


 
    # # Not Started NC SPRs owned by s/w team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    for sw_member in sw_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Submitted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Accepted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=In Review\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Postponed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_Not Started_NC'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # Not Started No DRB NC SPRs owned by s/w team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    for sw_member in sw_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Submitted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=No DRB\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Accepted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=No DRB\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=In Review\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=No DRB\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Postponed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=No DRB\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_Not Started_NC_No_DRB'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # Not Started Major NC SPRs owned by s/w team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    for sw_member in sw_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Submitted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Major\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Accepted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Major\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=In Review\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Major\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Postponed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Major\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_Not Started_NC_Major'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # Not Started Minor NC SPRs owned by s/w team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    for sw_member in sw_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Submitted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Minor\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Accepted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Minor\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=In Review\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Minor\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Postponed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Minor\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_Not Started_NC_Minor'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # # Not Started IO SPRs owned by s/w team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    for sw_member in sw_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Submitted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Accepted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=In Review\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Postponed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_Not Started_IO'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # in Prog SPRs owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000, \"=In Progress\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_in Prog'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # # in Prog NC owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000, \"=In Progress\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_in Prog_NC'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # # in Prog NC No_DRB owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000, \"=In Progress\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=No_DRB\")".format(ws_src.title) + "+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_in Prog_NC_No_DRB'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # # in Prog NC Major owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000, \"=In Progress\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Major\")".format(ws_src.title) + "+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_in Prog_NC_Major'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # in Prog NC Minor owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000, \"=In Progress\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Minor\")".format(ws_src.title) + "+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_in Prog_NC_Minor'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # in Prog IO owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000, \"=In Progress\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_in Prog_IO'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # Resolved SPRs owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000, \"=Resolved\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_Resolved'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # Verified SPRs owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000, \"=Verified\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_Verified'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # Closed SPRs owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000, \"=Closed\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_Closed'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # total SPRs owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ")+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_total'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # # NC SPRs owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_total_NC'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # NC NO DRB SPRs owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=No_DRB\")".format(ws_src.title) + "+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_total_NC_No_DRB'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # NC Major SPRs owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",'{0}'!AC2:AC3000,\"=Major\")".format(ws_src.title) + "+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_total_NC_Major'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # NC Minor SPRs owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",'{0}'!AC2:AC3000,\"=Minor\")".format(ws_src.title) + "+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_total_NC_Minor'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # IO  SPRs owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_total_IO'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # # Not Started SPRs owned by system team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    for system_member in system_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Submitted\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Accepted\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=In Review\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Postponed\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_Not Started'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


 
    # # Not Started NC SPRs owned by system team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    for system_member in system_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Submitted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Accepted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=In Review\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Postponed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_Not Started_NC'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # Not Started No DRB NC SPRs owned by system team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    for system_member in system_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Submitted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=No DRB\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Accepted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=No DRB\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=In Review\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=No DRB\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Postponed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=No DRB\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_Not Started_NC_No_DRB'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # Not Started Major NC SPRs owned by system team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    for system_member in system_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Submitted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Major\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Accepted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Major\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=In Review\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Major\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Postponed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Major\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_Not Started_NC_Major'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # Not Started Minor NC SPRs owned by system team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    for system_member in system_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Submitted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Minor\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Accepted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Minor\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=In Review\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Minor\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Postponed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Minor\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_Not Started_NC_Minor'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # # Not Started IO SPRs owned by system team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    for system_member in system_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Submitted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Accepted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=In Review\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Postponed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_Not Started_IO'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # in Prog SPRs owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000, \"=In Progress\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_in Prog'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # # in Prog NC owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000, \"=In Progress\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_in Prog_NC'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # # in Prog NC No_DRB owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000, \"=In Progress\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=No_DRB\")".format(ws_src.title) + "+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_in Prog_NC_No_DRB'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)


    # # in Prog NC Major owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000, \"=In Progress\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Major\")".format(ws_src.title) + "+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_in Prog_NC_Major'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # in Prog NC Minor owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000, \"=In Progress\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\"".format(ws_src.title) + ",{0}!AC2:AC3000,\"=Minor\")".format(ws_src.title) + "+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_in Prog_NC_Minor'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # in Prog IO owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000, \"=In Progress\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_in Prog_IO'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # Resolved SPRs owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000, \"=Resolved\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_Resolved'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # Verified SPRs owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000, \"=Verified\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_Verified'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    # # Closed SPRs owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000, \"=Closed\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_Closed'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position, index)

    index += 1

wb_src.save(file_name)
wb_src.close()
